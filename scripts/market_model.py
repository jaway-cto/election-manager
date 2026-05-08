"""
Market-calibrated per-council probability model.

For each council, outputs P(each party wins most seats) combining:
  1. PollCheck per-council seat ranges (FPTP-aware MRP base)
  2. Polymarket implied national totals (calibration target — "wisdom of money")
  3. Live declared seats (ground truth for Complete/Partial)

Run as part of poll_declared, or standalone:
    python market_model.py
"""

from __future__ import annotations

import json
import random
from collections import Counter
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from xlsx_lock import workbook_lock, save_with_retry

ROOT = Path(__file__).parent
XLSX = ROOT / "council_tracker_2026-05-08.xlsx"
PARTIES = ["Con", "Lab", "LD", "Grn", "Ref", "SNP", "PC", "Ind", "Oth"]
PRE_C, LAST_C, PREDICT_C, DECL_C = 8, 17, 26, 35

# Market-implied totals derived from Polymarket threshold ladders (50% point of YES probabilities).
# Update if Polymarket prices move materially.
# Polymarket-implied expected totals derived from threshold ladders.
# Recomputed 16:48 8 May 2026: Reform 1600+ has cooled to 32% (was 57%) — implied
# E[Reform] ≈ 1590. Other parties' upper thresholds aren't priced, so estimates
# are crude. Update these as Polymarket prices move materially.
MARKET_IMPLIED = {
    "Con": 740,
    "Lab": 880,
    "LD": 700,
    "Grn": 590,
    "Ref": 1590,
    "Ind": 90,
    "Oth": 130,
    "SNP": 0,
    "PC": 0,
}
# Shared "Reform momentum" SD as fraction of seats_up (within-council correlation factor).
# Higher -> wider scenarios, more uncertainty in pending councils.
MOMENTUM_SIGMA = 0.04  # 4 percentage points of seats moving between Ref and (Con+Lab)
MC_TRIALS = 4000


def gather_data(wb) -> tuple[list[dict], list[float]]:
    """Build per-council records. Returns (councils, market_implied_array)."""
    pollcheck = json.loads((ROOT / "pollcheck.json").read_text(encoding="utf-8")) if (ROOT / "pollcheck.json").exists() else {}

    councils: list[dict] = []
    for sheet in ("England", "Wales"):
        ws = wb[sheet]
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if not name or name == "TOTAL":
                continue
            seats_up = ws.cell(row=r, column=4).value or 0
            if not seats_up:
                continue
            status = ws.cell(row=r, column=6).value or "Pending"
            ctype = ws.cell(row=r, column=2).value
            subr = ws.cell(row=r, column=3).value
            last = [ws.cell(row=r, column=LAST_C + i).value or 0 for i in range(9)]
            decl = [ws.cell(row=r, column=DECL_C + i).value or 0 for i in range(9)]
            pc = pollcheck.get(name)
            councils.append({
                "name": name, "region": sheet, "type": ctype, "subregion": subr,
                "seats_up": seats_up, "status": status,
                "last": last, "declared": decl,
                "pollcheck": pc,
            })

    market_arr = [MARKET_IMPLIED[p] for p in PARTIES]
    return councils, market_arr


def compute_calibration(councils: list[dict], market: list[float]) -> list[float]:
    """For each party, scale = market_implied / sum_pollcheck_centrals_undeclared.
    Where PollCheck is missing, use last-election share as proxy."""
    pollcheck_total = [0.0] * 9
    for c in councils:
        if c["status"] in ("Complete", "Verified"):
            continue
        d_sum = sum(c["declared"])
        remaining = c["seats_up"] - d_sum
        if remaining <= 0:
            continue
        pc = c["pollcheck"]
        if pc and pc.get("projection"):
            pc_arr = [0] * 9
            party_idx = {p: i for i, p in enumerate(PARTIES)}
            for row in pc["projection"]:
                pc_arr[party_idx.get(row["party"], 8)] += row["central"]
            ps = sum(pc_arr) or 1
            for i in range(9):
                pollcheck_total[i] += pc_arr[i] * remaining / ps
        else:
            l_sum = sum(c["last"]) or 1
            for i in range(9):
                pollcheck_total[i] += c["last"][i] * remaining / l_sum

    declared_total = [0.0] * 9
    for c in councils:
        for i in range(9):
            declared_total[i] += c["declared"][i]

    # Reform-only calibration via bisection on aggregate after per-council renorm.
    # A naive 1.44 scale isn't enough: per-council renorm shrinks all parties when
    # totals exceed seats_up, eating most of Reform's gain. Iterate until aggregate
    # Reform matches market.
    ref_idx = PARTIES.index("Ref")
    target_ref = max(0.0, market[ref_idx] - declared_total[ref_idx])

    def aggregate_reform_at(s_ref: float) -> float:
        total = 0.0
        for c in councils:
            if c["status"] in ("Complete", "Verified"): continue
            d_sum = sum(c["declared"])
            remaining = c["seats_up"] - d_sum
            if remaining <= 0: continue
            pc = c["pollcheck"]
            party_idx = {p: i for i, p in enumerate(PARTIES)}
            if pc and pc.get("projection"):
                arr = [0.0] * 9
                for row in pc["projection"]:
                    arr[party_idx.get(row["party"], 8)] += row["central"]
            else:
                l_sum = sum(c["last"]) or 1.0
                arr = [c["last"][i] / l_sum * remaining for i in range(9)]
            arr[ref_idx] *= s_ref
            ssum = sum(arr) or 1.0
            total += arr[ref_idx] * remaining / ssum
        return total

    lo, hi = 1.0, 6.0
    for _ in range(30):
        mid = (lo + hi) / 2
        if aggregate_reform_at(mid) < target_ref:
            lo = mid
        else:
            hi = mid
        if hi - lo < 0.01: break
    scale = [1.0] * 9
    scale[ref_idx] = (lo + hi) / 2
    return scale


def project_council_remaining(c: dict, scale: list[float]) -> tuple[list[float], list[float]]:
    """Return (mean, sigma) per party for the council's REMAINING seats."""
    seats_up = c["seats_up"]
    d_sum = sum(c["declared"])
    remaining = max(seats_up - d_sum, 0)
    if remaining == 0:
        return [0.0] * 9, [0.0] * 9

    pc = c["pollcheck"]
    party_idx = {p: i for i, p in enumerate(PARTIES)}
    if pc and pc.get("projection"):
        pc_central = [0.0] * 9
        pc_low = [0.0] * 9
        pc_high = [0.0] * 9
        for row in pc["projection"]:
            idx = party_idx.get(row["party"], 8)
            pc_central[idx] += row["central"]
            pc_low[idx] += row["low"]
            pc_high[idx] += row["high"]
        # Apply calibration scaling
        adjusted = [pc_central[i] * scale[i] for i in range(9)]
        # Renormalise to fit remaining
        total = sum(adjusted) or 1.0
        means = [a * remaining / total for a in adjusted]
        # σ from PollCheck range (treat as ~80% CI, σ ≈ (high-low)/2.56)
        sigmas = [max(0.5, (pc_high[i] - pc_low[i]) / 2.56) for i in range(9)]
    else:
        # Fallback: last-election share scaled
        l_sum = sum(c["last"]) or 1.0
        adjusted = [c["last"][i] / l_sum * scale[i] for i in range(9)]
        total = sum(adjusted) or 1.0
        means = [a * remaining / total for a in adjusted]
        # σ wider when no PollCheck data
        sigmas = [max(1.0, remaining * 0.10) for _ in range(9)]
    return means, sigmas


def monte_carlo(c: dict, scale: list[float]) -> dict:
    """Sample seat allocations and count argmax frequency."""
    if c["status"] in ("Complete", "Verified") and sum(c["declared"]) > 0:
        seats = c["declared"]
        winner = max(range(9), key=lambda i: seats[i])
        probs = [0.0] * 9
        # 99% on declared winner, 1% spread over others (verification fudge)
        probs[winner] = 0.99
        for i in range(9):
            if i != winner and seats[i] > 0:
                probs[i] = 0.01 / max(1, sum(1 for s in seats if s > 0) - 1)
        runner_i = sorted(range(9), key=lambda i: -seats[i])[1]
        return {
            "probs": probs, "expected_seats": [float(s) for s in seats],
            "winner_modal": PARTIES[winner], "winner_prob": probs[winner],
            "runner_modal": PARTIES[runner_i],
        }

    means, sigmas = project_council_remaining(c, scale)
    declared = c["declared"]
    seats_up = c["seats_up"]

    win_counts = Counter()
    expected_seats = [0.0] * 9
    for _ in range(MC_TRIALS):
        # Per-party noise + shared Reform-momentum
        momentum = random.gauss(0, MOMENTUM_SIGMA * seats_up)
        sample = []
        for i in range(9):
            base = means[i]
            noise = random.gauss(0, sigmas[i])
            if PARTIES[i] == "Ref":
                base += momentum
            elif PARTIES[i] in ("Con", "Lab"):
                base -= momentum / 2  # split the momentum across two main losers
            sample.append(max(0.0, base + noise))
        # Add declared seats
        total = [declared[i] + sample[i] for i in range(9)]
        # Renormalise total so it equals seats_up (preserves council size)
        ssum = sum(total) or 1.0
        total = [t * seats_up / ssum for t in total]
        for i in range(9):
            expected_seats[i] += total[i]
        winner = max(range(9), key=lambda i: total[i])
        win_counts[winner] += 1

    probs = [win_counts[i] / MC_TRIALS for i in range(9)]
    expected_seats = [s / MC_TRIALS for s in expected_seats]
    winner_i = max(range(9), key=lambda i: probs[i])
    runner_i = sorted(range(9), key=lambda i: -probs[i])[1]
    return {
        "probs": probs, "expected_seats": expected_seats,
        "winner_modal": PARTIES[winner_i], "winner_prob": probs[winner_i],
        "runner_modal": PARTIES[runner_i],
    }


def write_tab(wb, councils: list[dict], scale: list[float], results: dict[str, dict]) -> None:
    name = "Market model"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ws["A1"] = "Market-calibrated per-council probabilities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Calibration: PollCheck centrals scaled to Polymarket-implied totals.  "
                f"Scaling factors: " + " ".join(f"{p}×{s:.2f}" for p, s in zip(PARTIES, scale) if 0 < s < 5 and MARKET_IMPLIED[p] > 0))
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = (f"Method: Monte Carlo (n={MC_TRIALS}) per pending council using PollCheck low-high "
                f"as σ, plus shared Reform-momentum factor (σ={MOMENTUM_SIGMA*100:.1f}pp seats_up).  "
                f"Complete councils: deterministic 99% on declared winner.")
    ws["A3"].font = Font(italic=True, color="666666")

    headers = ["Region", "Council", "Type", "Status", "Seats up",
               "Modal winner", "Win P", "Runner-up", "Runner P",
               "BBC winner", "BBC flash"] + \
              [f"P({p})" for p in PARTIES] + \
              [f"E[{p}]" for p in PARTIES]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")

    # Load BBC per-council winners
    bbc_path = ROOT / "bbc_data.json"
    bbc_winners: dict[str, dict] = {}
    if bbc_path.exists():
        bbc_data = json.loads(bbc_path.read_text(encoding="utf-8"))
        for cw in bbc_data.get("councils", []):
            bbc_winners[cw["name"]] = cw

    def match_bbc(name: str) -> dict | None:
        # Direct match first
        if name in bbc_winners:
            return bbc_winners[name]
        # Fuzzy: strip suffixes and try
        normalised = (name.replace(" CC", "").replace(" (Mayor)", "")
                      .replace(" (by-election)", "").replace(" (shadow)", "")
                      .replace("Hull (Kingston upon Hull)", "Kingston upon Hull")
                      .strip())
        for k in bbc_winners:
            if k == normalised or k.replace("&", "and") == normalised.replace("&", "and"):
                return bbc_winners[k]
        return None

    row = 6
    for c in councils:
        res = results[c["name"]]
        ws.cell(row=row, column=1, value=c["region"])
        ws.cell(row=row, column=2, value=c["name"])
        ws.cell(row=row, column=3, value=c["type"])
        ws.cell(row=row, column=4, value=c["status"])
        ws.cell(row=row, column=5, value=c["seats_up"])
        ws.cell(row=row, column=6, value=res["winner_modal"])
        ws.cell(row=row, column=7, value=res["winner_prob"]).number_format = "0.0%"
        ws.cell(row=row, column=8, value=res["runner_modal"])
        runner_i = PARTIES.index(res["runner_modal"])
        ws.cell(row=row, column=9, value=res["probs"][runner_i]).number_format = "0.0%"
        # BBC columns
        bbc = match_bbc(c["name"])
        if bbc and bbc.get("declared"):
            ws.cell(row=row, column=10, value=bbc.get("winner_party") or "")
            ws.cell(row=row, column=11, value=bbc.get("flash") or "")
            # Highlight model/BBC mismatch
            if bbc.get("winner_party") and bbc["winner_party"] not in (res["winner_modal"], "NOC"):
                ws.cell(row=row, column=10).fill = PatternFill("solid", fgColor="FFC7CE")
        for i, p in enumerate(PARTIES):
            ws.cell(row=row, column=12 + i, value=res["probs"][i]).number_format = "0.0%"
            ws.cell(row=row, column=21 + i, value=round(res["expected_seats"][i], 1))
        row += 1

    # Aggregate row at bottom (E columns now start at 21 due to added BBC columns)
    ws.cell(row=row + 1, column=1, value="Aggregate expected seats (sense-check vs market)").font = Font(bold=True)
    for i in range(9):
        total = sum(results[c["name"]]["expected_seats"][i] for c in councils)
        ws.cell(row=row + 1, column=21 + i, value=round(total)).font = Font(bold=True)
    ws.cell(row=row + 2, column=1, value="Polymarket-implied totals").font = Font(italic=True)
    for i, p in enumerate(PARTIES):
        ws.cell(row=row + 2, column=21 + i, value=MARKET_IMPLIED[p]).font = Font(italic=True)

    widths = {1: 11, 2: 36, 3: 14, 4: 12, 5: 9, 6: 14, 7: 9, 8: 12, 9: 9, 10: 12, 11: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    for col in range(12, 30):
        ws.column_dimensions[get_column_letter(col)].width = 7
    ws.freeze_panes = "F6"


def run() -> None:
    with workbook_lock():
        wb = openpyxl.load_workbook(XLSX)
        councils, market = gather_data(wb)
        scale = compute_calibration(councils, market)
        results = {c["name"]: monte_carlo(c, scale) for c in councils}
        write_tab(wb, councils, scale, results)
        save_with_retry(wb, XLSX)
    print(f"Market model wrote {len(councils)} councils. Calibration scales:",
          {p: round(s, 2) for p, s in zip(PARTIES, scale) if MARKET_IMPLIED[p] > 0})


if __name__ == "__main__":
    run()
