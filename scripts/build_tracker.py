"""
Council elections live tracker — 2026-05-08
Builds an .xlsx with: Summary | England | Scotland | Wales tabs.

Re-run this script to rebuild the empty skeleton. Council rows are
populated from COUNCILS_* lists below (filled in once research agents return).
Live data entry happens directly in Excel — formulas + validation drive the rest.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule

OUT = r"C:\Users\ringh\elections-2026\council_tracker_2026-05-08.xlsx"

# Party columns — superset across all 3 regions; zeros where party didn't stand/win.
PARTIES = ["Con", "Lab", "LD", "Grn", "Ref", "SNP", "PC", "Ind", "Oth"]
N = len(PARTIES)

from councils_data import ENGLAND as COUNCILS_ENGLAND, SCOTLAND as COUNCILS_SCOTLAND, WALES as COUNCILS_WALES

def _pad(arr):
    """councils_data.py stores arrays in legacy 7-party order
    (Con,Lab,LD,Grn,Ref,Ind,Oth). Pad to 9 by inserting SNP=PC=0 at index 5."""
    if len(arr) == N:
        return list(arr)
    if len(arr) == 7:
        return list(arr[:5]) + [0, 0] + list(arr[5:])
    raise ValueError(f"Unexpected party array length {len(arr)}: {arr}")

import json
from pathlib import Path
_LAST_JSON = Path(__file__).parent / "last_seats.json"
_LAST_OVERRIDES = json.loads(_LAST_JSON.read_text(encoding="utf-8")) if _LAST_JSON.exists() else {}

def _norm(councils):
    out = []
    for c in councils:
        name = c[0]
        pre = _pad(c[5])
        last_existing = _pad(c[6])
        # Override with HoC-derived seats unless councils_data.py supplies non-zero values
        last = _LAST_OVERRIDES.get(name) if sum(last_existing) == 0 else last_existing
        if last is None:
            last = last_existing
        out.append((name, c[1], c[2], c[3], c[4], pre, last, c[7]))
    return out

COUNCILS_ENGLAND = _norm(COUNCILS_ENGLAND)
COUNCILS_SCOTLAND = _norm(COUNCILS_SCOTLAND)
COUNCILS_WALES = _norm(COUNCILS_WALES)

STATUSES = ["Pending", "Counting", "Partial", "Complete", "Verified"]

# ---- Styling ----
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHDR_FILL = PatternFill("solid", fgColor="D9E1F2")
SUBHDR_FONT = Font(bold=True, size=10)
SECTION_FILLS = {
    "PRE":      PatternFill("solid", fgColor="FFF2CC"),  # cream
    "LAST":     PatternFill("solid", fgColor="E2EFDA"),  # green
    "PREDICT":  PatternFill("solid", fgColor="FCE4D6"),  # peach
    "DECLARED": PatternFill("solid", fgColor="DDEBF7"),  # blue
}
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Column layout (1-indexed).
# A Council | B Type | C Sub-region | D Seats Up | E Total Seats |
# F Status | G Declared At |
# H..N PRE (7) | O..U LAST (7) | V..AB PREDICT (7) | AC..AI DECLARED (7) |
# AJ Declared Sum | AK Control Before | AL Control After | AM Net headline | AN Notes
# Column layout (1-indexed). Block width = N parties.
_PRE = 8
COL = {
    "council": 1, "type": 2, "subregion": 3, "seats_up": 4, "total_seats": 5,
    "status": 6, "declared_at": 7,
    "pre_start":     _PRE,
    "last_start":    _PRE + N,
    "predict_start": _PRE + 2*N,
    "decl_start":    _PRE + 3*N,
    "decl_sum":      _PRE + 4*N,
    "ctrl_before":   _PRE + 4*N + 1,
    "ctrl_after":    _PRE + 4*N + 2,
    "net":           _PRE + 4*N + 3,
    "notes":         _PRE + 4*N + 4,
}
LAST_COL = COL["notes"]


def write_region_sheet(wb: Workbook, name: str, councils: list[tuple]) -> None:
    ws = wb.create_sheet(name)

    # Row 1: section banners
    sections = [
        ("Council info", COL["council"], COL["declared_at"], None),
        ("Pre-election seats (current council)", COL["pre_start"], COL["pre_start"] + N - 1, "PRE"),
        ("Last comparable election (seats won)", COL["last_start"], COL["last_start"] + N - 1, "LAST"),
        ("Prediction (seats)", COL["predict_start"], COL["predict_start"] + N - 1, "PREDICT"),
        ("Declared today (seats won)", COL["decl_start"], COL["decl_start"] + N - 1, "DECLARED"),
        ("Roll-ups", COL["decl_sum"], COL["notes"], None),
    ]
    for label, c1, c2, key in sections:
        ws.cell(row=1, column=c1, value=label).font = HDR_FONT
        ws.cell(row=1, column=c1).fill = HDR_FILL if key is None else SECTION_FILLS[key]
        ws.cell(row=1, column=c1).alignment = Alignment(horizontal="center")
        if c2 > c1:
            ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)

    # Row 2: column headers
    headers = {
        COL["council"]: "Council",
        COL["type"]: "Type",
        COL["subregion"]: "Sub-region",
        COL["seats_up"]: "Seats Up",
        COL["total_seats"]: "Total Council Seats",
        COL["status"]: "Status",
        COL["declared_at"]: "Declared At",
        COL["decl_sum"]: "Declared Sum",
        COL["ctrl_before"]: "Control Before",
        COL["ctrl_after"]: "Control After",
        COL["net"]: "Net (vs last)",
        COL["notes"]: "Notes / source",
    }
    for c, h in headers.items():
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = SUBHDR_FONT
        cell.fill = SUBHDR_FILL
    for block_key, start in [("PRE", COL["pre_start"]), ("LAST", COL["last_start"]),
                             ("PREDICT", COL["predict_start"]), ("DECLARED", COL["decl_start"])]:
        for i, p in enumerate(PARTIES):
            cell = ws.cell(row=2, column=start + i, value=p)
            cell.font = SUBHDR_FONT
            cell.fill = SECTION_FILLS[block_key]

    # Council rows
    DATA_START = 3
    row = DATA_START
    for c in councils:
        council, ctype, subr, seats_up, total_seats, pre, last, ctrl = c
        ws.cell(row=row, column=COL["council"], value=council).font = Font(bold=True)
        ws.cell(row=row, column=COL["type"], value=ctype)
        ws.cell(row=row, column=COL["subregion"], value=subr)
        ws.cell(row=row, column=COL["seats_up"], value=seats_up)
        ws.cell(row=row, column=COL["total_seats"], value=total_seats)
        ws.cell(row=row, column=COL["status"], value="Pending")
        for i, v in enumerate(pre):
            ws.cell(row=row, column=COL["pre_start"] + i, value=v)
        for i, v in enumerate(last):
            ws.cell(row=row, column=COL["last_start"] + i, value=v)
        # predict + declared left blank for manual entry
        ws.cell(row=row, column=COL["ctrl_before"], value=ctrl)
        # Declared sum formula
        c1 = get_column_letter(COL["decl_start"])
        c2 = get_column_letter(COL["decl_start"] + N - 1)
        ws.cell(row=row, column=COL["decl_sum"],
                value=f"=SUM({c1}{row}:{c2}{row})")
        # Net headline: declared sum minus last election sum (text not necessary; numeric helps summary)
        last_c1 = get_column_letter(COL["last_start"])
        last_c2 = get_column_letter(COL["last_start"] + N - 1)
        # leave Net blank — computed in summary per party. (Per-council headline is hard to summarise textually.)
        row += 1

    # Reserve 60 blank rows beyond actual data so users can add councils without breaking validation
    BLANK_TAIL = 60
    last_data_row = row - 1
    validation_last_row = last_data_row + BLANK_TAIL

    # Add empty status cells with default "Pending" for blank rows so dropdown is visible
    for r in range(row, row + BLANK_TAIL):
        ws.cell(row=r, column=COL["status"], value=None)

    # Data validation: status dropdown
    dv = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=True)
    dv.add(f"{get_column_letter(COL['status'])}3:{get_column_letter(COL['status'])}{validation_last_row}")
    ws.add_data_validation(dv)

    # Conditional formatting on status
    status_col = get_column_letter(COL["status"])
    rng = f"{status_col}3:{status_col}{validation_last_row}"
    rules = [
        ("Pending",  "BFBFBF"),
        ("Counting", "FFE699"),
        ("Partial",  "F4B084"),
        ("Complete", "A9D08E"),
        ("Verified", "548235"),
    ]
    for txt, colr in rules:
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=[f'"{txt}"'],
                       fill=PatternFill("solid", fgColor=colr)))

    # Column widths
    widths = {1: 32, 2: 14, 3: 18, 4: 9, 5: 12, 6: 12, 7: 14,
              COL["decl_sum"]: 12, COL["ctrl_before"]: 14, COL["ctrl_after"]: 14,
              COL["net"]: 14, COL["notes"]: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    # Party block columns narrow
    for block_start in [COL["pre_start"], COL["last_start"], COL["predict_start"], COL["decl_start"]]:
        for i in range(N):
            ws.column_dimensions[get_column_letter(block_start + i)].width = 6

    ws.freeze_panes = "B3"

    # Totals row at top of data area? Put a TOTAL row just under headers (row 3 reserved? no - keep data starting row 3, totals at bottom)
    total_row = validation_last_row + 2
    ws.cell(row=total_row, column=COL["council"], value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=COL["seats_up"],
            value=f"=SUM({get_column_letter(COL['seats_up'])}3:{get_column_letter(COL['seats_up'])}{validation_last_row})").font = Font(bold=True)
    for block_start in [COL["pre_start"], COL["last_start"], COL["predict_start"], COL["decl_start"]]:
        for i in range(N):
            col_l = get_column_letter(block_start + i)
            ws.cell(row=total_row, column=block_start + i,
                    value=f"=SUM({col_l}3:{col_l}{validation_last_row})").font = Font(bold=True)
    ws.cell(row=total_row, column=COL["decl_sum"],
            value=f"=SUM({get_column_letter(COL['decl_sum'])}3:{get_column_letter(COL['decl_sum'])}{validation_last_row})").font = Font(bold=True)


def write_summary(wb: Workbook) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws["A1"] = "UK Council Elections — Live Tracker"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = "Polling day: 2026-05-07  |  Declaration day: 2026-05-08"
    ws["A2"].font = Font(italic=True, color="666666")

    regions = ["England", "Scotland", "Wales"]

    # --- Headline block (row 4) ---
    ws["A4"] = "Headline"
    ws["A4"].font = HDR_FONT
    ws["A4"].fill = HDR_FILL
    ws.merge_cells("A4:G4")

    headers = ["Region", "Councils Up", "Declared", "% Declared", "Seats Up", "Seats Declared", "% Seats Declared"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = SUBHDR_FONT
        c.fill = SUBHDR_FILL

    for ridx, region in enumerate(regions, start=6):
        ws.cell(row=ridx, column=1, value=region).font = Font(bold=True)
        # Councils up = COUNTA of council column on region tab
        ws.cell(row=ridx, column=2, value=f"=COUNTA('{region}'!A3:A1000)-COUNTIF('{region}'!A3:A1000,\"TOTAL\")")
        # Declared councils = COUNTIF status = Complete or Verified
        ws.cell(row=ridx, column=3, value=f"=COUNTIF('{region}'!F3:F1000,\"Complete\")+COUNTIF('{region}'!F3:F1000,\"Verified\")")
        ws.cell(row=ridx, column=4, value=f"=IFERROR(C{ridx}/B{ridx},0)").number_format = "0.0%"
        # Seats up = sum of seats up column
        ws.cell(row=ridx, column=5, value=f"=SUMIF('{region}'!F3:F1000,\"<>\",'{region}'!D3:D1000)")
        # Seats declared = sum of declared sum column where status complete/verified
        ws.cell(row=ridx, column=6,
                value=(f"=SUMIFS('{region}'!{get_column_letter(COL['decl_sum'])}3:{get_column_letter(COL['decl_sum'])}1000,'{region}'!F3:F1000,\"Complete\")"
                       f"+SUMIFS('{region}'!{get_column_letter(COL['decl_sum'])}3:{get_column_letter(COL['decl_sum'])}1000,'{region}'!F3:F1000,\"Verified\")"))
        ws.cell(row=ridx, column=7, value=f"=IFERROR(F{ridx}/E{ridx},0)").number_format = "0.0%"

    # Total row
    tr = 9
    ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 8):
        col_l = get_column_letter(col)
        if col in (4, 7):
            ws.cell(row=tr, column=col, value=f"=IFERROR({get_column_letter(col-1)}{tr}/{get_column_letter(col-2)}{tr},0)").number_format = "0.0%"
        else:
            ws.cell(row=tr, column=col, value=f"=SUM({col_l}6:{col_l}8)").font = Font(bold=True)

    # --- Per-party totals block ---
    start_row = 12
    ws.cell(row=start_row, column=1, value="Party totals (all regions)").font = HDR_FONT
    ws.cell(row=start_row, column=1).fill = HDR_FILL
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)

    party_headers = ["Party", "Pre-election seats", "Last comparable", "Predicted",
                     "Declared so far", "Net vs Pre", "Net vs Last", "% of decl seats"]
    for i, h in enumerate(party_headers, 1):
        c = ws.cell(row=start_row + 1, column=i, value=h)
        c.font = SUBHDR_FONT
        c.fill = SUBHDR_FILL

    # Range columns on region tabs
    BLOCKS = {
        "pre":     (get_column_letter(COL["pre_start"]),     get_column_letter(COL["pre_start"]     + N - 1)),
        "last":    (get_column_letter(COL["last_start"]),    get_column_letter(COL["last_start"]    + N - 1)),
        "predict": (get_column_letter(COL["predict_start"]), get_column_letter(COL["predict_start"] + N - 1)),
        "decl":    (get_column_letter(COL["decl_start"]),    get_column_letter(COL["decl_start"]    + N - 1)),
    }

    for i, party in enumerate(PARTIES):
        r = start_row + 2 + i
        ws.cell(row=r, column=1, value=party).font = Font(bold=True)
        # Each party occupies offset i within its block
        # PRE: column = pre_start + i (H + i)
        def col_for(block_key: str) -> str:
            start_letter = BLOCKS[block_key][0]
            # convert letter to number, add i
            from openpyxl.utils import column_index_from_string
            return get_column_letter(column_index_from_string(start_letter) + i)

        pre_c = col_for("pre"); last_c = col_for("last")
        pred_c = col_for("predict"); decl_c = col_for("decl")

        # Sum across all 3 region tabs
        def sum_across(col_letter: str) -> str:
            parts = [f"SUM('{rg}'!{col_letter}3:{col_letter}1000)" for rg in regions]
            # Subtract the TOTAL row contribution (TOTAL row contains a sum that would double-count if row<=1000).
            # Our region totals sit at row >> 1000-ish? actually we placed at validation_last_row+2 which is small.
            # To avoid double-counting, exclude TOTAL row by filtering on column A<>"TOTAL":
            parts = [f"SUMIF('{rg}'!A3:A1000,\"<>TOTAL\",'{rg}'!{col_letter}3:{col_letter}1000)" for rg in regions]
            return "=" + "+".join(parts)

        ws.cell(row=r, column=2, value=sum_across(pre_c))
        ws.cell(row=r, column=3, value=sum_across(last_c))
        ws.cell(row=r, column=4, value=sum_across(pred_c))
        ws.cell(row=r, column=5, value=sum_across(decl_c))
        ws.cell(row=r, column=6, value=f"=E{r}-B{r}")
        ws.cell(row=r, column=7, value=f"=E{r}-C{r}")
        ws.cell(row=r, column=8, value=f"=IFERROR(E{r}/SUM(E{start_row+2}:E{start_row+2+len(PARTIES)-1}),0)").number_format = "0.0%"

    # Party total row
    ptot = start_row + 2 + len(PARTIES)
    ws.cell(row=ptot, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 8):
        col_l = get_column_letter(col)
        ws.cell(row=ptot, column=col, value=f"=SUM({col_l}{start_row+2}:{col_l}{ptot-1})").font = Font(bold=True)

    # --- Per-region per-party block ---
    next_row = ptot + 3
    for region in regions:
        ws.cell(row=next_row, column=1, value=f"{region} — party breakdown").font = HDR_FONT
        ws.cell(row=next_row, column=1).fill = HDR_FILL
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=8)

        for i, h in enumerate(party_headers, 1):
            c = ws.cell(row=next_row + 1, column=i, value=h)
            c.font = SUBHDR_FONT
            c.fill = SUBHDR_FILL

        for i, party in enumerate(PARTIES):
            r = next_row + 2 + i
            ws.cell(row=r, column=1, value=party).font = Font(bold=True)
            from openpyxl.utils import column_index_from_string
            def col_for(block_key: str) -> str:
                start_letter = BLOCKS[block_key][0]
                return get_column_letter(column_index_from_string(start_letter) + i)
            pre_c = col_for("pre"); last_c = col_for("last")
            pred_c = col_for("predict"); decl_c = col_for("decl")
            ws.cell(row=r, column=2, value=f"=SUMIF('{region}'!A3:A1000,\"<>TOTAL\",'{region}'!{pre_c}3:{pre_c}1000)")
            ws.cell(row=r, column=3, value=f"=SUMIF('{region}'!A3:A1000,\"<>TOTAL\",'{region}'!{last_c}3:{last_c}1000)")
            ws.cell(row=r, column=4, value=f"=SUMIF('{region}'!A3:A1000,\"<>TOTAL\",'{region}'!{pred_c}3:{pred_c}1000)")
            ws.cell(row=r, column=5, value=f"=SUMIF('{region}'!A3:A1000,\"<>TOTAL\",'{region}'!{decl_c}3:{decl_c}1000)")
            ws.cell(row=r, column=6, value=f"=E{r}-B{r}")
            ws.cell(row=r, column=7, value=f"=E{r}-C{r}")
            ws.cell(row=r, column=8, value=f"=IFERROR(E{r}/SUM(E{next_row+2}:E{next_row+2+len(PARTIES)-1}),0)").number_format = "0.0%"
        next_row += len(PARTIES) + 4

    # --- Projection block ---
    # Per-party live projection. Cells are populated by poll_declared.py (Python computation,
    # exact). Earlier attempt at array SUMPRODUCT formulas was broken — IFERROR collapses
    # ranges to scalars on div-by-zero. Static-write approach is accurate and simpler.
    proj_row = next_row
    ws.cell(row=proj_row, column=1, value="Live projection — implied total per party (refreshed by poller)").font = HDR_FONT
    ws.cell(row=proj_row, column=1).fill = HDR_FILL
    ws.merge_cells(start_row=proj_row, start_column=1, end_row=proj_row, end_column=8)

    proj_headers = ["Party", "Declared", "% of declared", "Same-as-last projected total",
                    "% of total seats", "PREDICT-based projected total", "% of total seats", ""]
    for i, h in enumerate(proj_headers, 1):
        c = ws.cell(row=proj_row + 1, column=i, value=h)
        c.font = SUBHDR_FONT
        c.fill = SUBHDR_FILL

    # Empty rows; poll_declared.update_projection_summary() writes static values.
    n_parties = len(PARTIES)
    for i, party in enumerate(PARTIES):
        r = proj_row + 2 + i
        ws.cell(row=r, column=1, value=party).font = Font(bold=True)
        for col in (4, 6):
            ws.cell(row=r, column=col).number_format = "0"
        for col in (3, 5, 7):
            ws.cell(row=r, column=col).number_format = "0.0%"
    ptot = proj_row + 2 + n_parties
    ws.cell(row=ptot, column=1, value="TOTAL").font = Font(bold=True)

    next_row = ptot + 3

    # --- Notes block ---
    ws.cell(row=next_row, column=1, value="Notes").font = Font(bold=True)
    notes = [
        "Status values: Pending → Counting → Partial → Complete → Verified.",
        "Only Complete and Verified rows count as 'declared' in headline.",
        "Party columns: Con, Lab, LD (Lib Dem), Grn (Green), Ref (Reform UK), Ind (Independent), Oth (any other — note in 'Notes' column).",
        "For Wales: enter Plaid Cymru seats under 'Oth' OR rename column locally; for Scotland: SNP under 'Oth' similarly. (See per-region tab for guidance.)",
        "PRE block = composition of council going into election (current seats per party).",
        "LAST block = seats won by each party at the last comparable election (the seats up THIS time).",
        "PREDICT block = your forecast — fill in before declarations start; compare vs DECLARED.",
        "DECLARED block = live results entered as wards declare. Set Status to Complete when full council declared.",
        "Declared Sum (column AJ) should equal Seats Up (column D) when council fully declared — quick sanity check.",
    ]
    for i, n in enumerate(notes):
        ws.cell(row=next_row + 1 + i, column=1, value="• " + n)
        ws.merge_cells(start_row=next_row + 1 + i, start_column=1, end_row=next_row + 1 + i, end_column=8)

    # Column widths
    ws.column_dimensions["A"].width = 28
    for col in range(2, 9):
        ws.column_dimensions[get_column_letter(col)].width = 16


def main() -> None:
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    write_region_sheet(wb, "England", COUNCILS_ENGLAND)
    write_region_sheet(wb, "Scotland", COUNCILS_SCOTLAND)
    write_region_sheet(wb, "Wales", COUNCILS_WALES)
    write_summary(wb)

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
