"""
Poll Polymarket Gamma API and write live odds + edge analysis into a Markets tab.

READ-ONLY: only HTTP GET to https://gamma-api.polymarket.com/events?slug=...
No authentication. No transactions. Cannot place bets.

Usage:
    python poll_markets.py                # one-shot
    python poll_markets.py --watch        # poll every 60s
    python poll_markets.py --watch --interval 120
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.request
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from xlsx_lock import workbook_lock, save_with_retry

ROOT = Path(__file__).parent
XLSX = ROOT / "council_tracker_2026-05-08.xlsx"

GAMMA = "https://gamma-api.polymarket.com/events?slug="
UA = {"User-Agent": "elections-tracker/1.0 (read-only odds fetch)"}

# (event_slug, party_label) -- party_label maps to our PARTIES if it's a seat-threshold event
EVENTS = [
    ("2026-united-kingdom-local-elections-reform-wins-seats",       "Ref"),
    ("2026-united-kingdom-local-elections-labour-wins-seats",       "Lab"),
    ("2026-united-kingdom-local-elections-conservative-wins-seats", "Con"),
    ("2026-united-kingdom-local-elections-green-wins-seats",        "Grn"),
    ("2026-united-kingdom-local-elections-party-winner",            None),
    ("2026-united-kingdom-local-elections-2nd-place",               None),
    ("will-reform-win-a-mayorship-in-the-2026-united-kingdom-local-elections", None),
]

PARTIES = ["Con", "Lab", "LD", "Grn", "Ref", "SNP", "PC", "Ind", "Oth"]

THRESHOLD_RE = re.compile(r"at least (\d+) council seat", re.I)


def fetch_event(slug: str) -> dict | None:
    req = urllib.request.Request(GAMMA + slug, headers=UA)
    try:
        with urllib.request.urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read())
    except Exception as e:
        print(f"  [{slug}] fetch error: {e}", file=sys.stderr)
        return None
    if not data:
        return None
    return data[0]


def parse_market(m: dict) -> dict | None:
    """Extract the fields we care about. Polymarket can return outcomes/prices as JSON strings."""
    if m.get("closed"):
        return None
    outs = m.get("outcomes")
    prices = m.get("outcomePrices")
    if outs is None or prices is None:
        return None
    if isinstance(outs, str):
        try: outs = json.loads(outs)
        except json.JSONDecodeError: return None
    if isinstance(prices, str):
        try: prices = json.loads(prices)
        except json.JSONDecodeError: return None
    if not prices:
        return None
    try:
        yes_price = float(prices[0])
    except (TypeError, ValueError):
        return None
    return {
        "id": m.get("id"),
        "question": m.get("question", ""),
        "yes_price": yes_price,
        "last_trade": m.get("lastTradePrice"),
        "volume": float(m.get("volume", 0) or 0),
        "volume_24h": float(m.get("volume24hr", 0) or 0),
        "active": m.get("active"),
    }


def fetch_all_markets() -> list[dict]:
    rows: list[dict] = []
    for slug, party in EVENTS:
        e = fetch_event(slug)
        if not e:
            continue
        for m in e.get("markets", []):
            parsed = parse_market(m)
            if not parsed or parsed["yes_price"] is None:
                continue
            # Skip dummy "Party A/B/C..." placeholder markets which have zero volume + identical names
            if parsed["volume"] == 0 and "Party " in parsed["question"]:
                continue
            threshold = None
            mt = THRESHOLD_RE.search(parsed["question"])
            if mt:
                threshold = int(mt.group(1))
            rows.append({
                "event_title": e.get("title", slug),
                "party": party,
                "threshold": threshold,
                **parsed,
            })
    return rows


def get_live_data() -> dict:
    """Return current declared seats per party + total declared/total-up from the workbook."""
    wb = openpyxl.load_workbook(XLSX, data_only=False)
    PRE, LAST, DECL = 8, 17, 35
    decl = [0] * 9
    last = [0] * 9
    seats_up = 0
    for sheet in ("England", "Wales"):
        ws = wb[sheet]
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if not name or name == "TOTAL":
                continue
            su = ws.cell(row=r, column=4).value or 0
            seats_up += su
            for i in range(9):
                last[i] += ws.cell(row=r, column=LAST + i).value or 0
                decl[i] += ws.cell(row=r, column=DECL + i).value or 0
    decl_total = sum(decl)
    return {
        "declared_per_party": decl,
        "last_per_party": last,
        "declared_total": decl_total,
        "seats_up_total": seats_up,
        "remaining_total": seats_up - decl_total,
    }


def projections(live: dict, party_label: str) -> dict:
    """Two crude extrapolations of final total seats for one party:
       - extrap_rate: declared_share * total_seats_up (assumes remaining = declared so far)
       - same_as_last: declared + last_party * remaining/last_total"""
    if party_label not in PARTIES:
        return {"extrap_rate": None, "same_as_last": None}
    i = PARTIES.index(party_label)
    decl_total = live["declared_total"]
    seats_up = live["seats_up_total"]
    last_total = sum(live["last_per_party"])
    remaining = live["remaining_total"]
    extrap = (live["declared_per_party"][i] / decl_total * seats_up) if decl_total > 0 else None
    sal = (live["declared_per_party"][i]
           + live["last_per_party"][i] * remaining / last_total) if last_total > 0 else None
    return {"extrap_rate": extrap, "same_as_last": sal}


def write_markets_tab(rows: list[dict], live: dict, dry_run: bool = False) -> None:
    with workbook_lock():
        _write_markets_inner(rows, live, dry_run)


def _write_markets_inner(rows: list[dict], live: dict, dry_run: bool) -> None:
    wb = openpyxl.load_workbook(XLSX)
    if "Markets" in wb.sheetnames:
        del wb["Markets"]
    ws = wb.create_sheet("Markets")

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A1"] = "Polymarket live odds — UK 2026 local elections"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Last refreshed: {ts}    Source: gamma-api.polymarket.com (read-only, no auth)"
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = (f"Live data: {live['declared_total']:,} of {live['seats_up_total']:,} seats "
                f"declared ({100*live['declared_total']/live['seats_up_total']:.1f}%)  |  "
                f"Remaining: {live['remaining_total']:,}")
    ws["A3"].font = Font(italic=True)

    headers = [
        "Event", "Party", "Threshold", "Question", "Volume ($)", "24h Vol ($)",
        "YES price", "YES last trade", "Implied probability",
        "Live declared (party)", "Extrapolate-rate proj", "Same-as-last proj",
        "Extrap clears threshold?", "Same-as-last clears?", "Edge flag",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")

    # Sort: party-threshold events first (by party then threshold), then party-winner / 2nd-place / mayorship
    def sort_key(r: dict):
        return (r["party"] or "Z", r["threshold"] if r["threshold"] is not None else 9999, r["question"])
    rows.sort(key=sort_key)

    row = 6
    for r in rows:
        proj = projections(live, r["party"]) if r["party"] else {"extrap_rate": None, "same_as_last": None}
        live_decl = live["declared_per_party"][PARTIES.index(r["party"])] if r["party"] in PARTIES else None

        # Edge flag: market-implied YES vs whether our projections clear threshold
        edge = ""
        if r["threshold"] is not None and proj["extrap_rate"] is not None:
            extrap_yes = 1 if proj["extrap_rate"] >= r["threshold"] else 0
            sal_yes = (1 if proj["same_as_last"] is not None
                       and proj["same_as_last"] >= r["threshold"] else 0)
            mp = r["yes_price"]
            # Only flag where we strongly disagree with market in either direction
            # Neutral wording — flags disagreement, not betting advice. The geographic
            # mix of remaining declarations may make either side correct.
            if extrap_yes and mp < 0.4:
                edge = f"DISAGREE: mkt {mp:.0%} NO, rate-extrap clears threshold"
            elif not extrap_yes and mp > 0.6:
                edge = f"DISAGREE: mkt {mp:.0%} YES, rate-extrap below threshold"

        cells = [
            r["event_title"], r["party"] or "", r["threshold"] or "", r["question"],
            r["volume"], r["volume_24h"],
            r["yes_price"], r["last_trade"],
            r["yes_price"],  # same as YES price for display (formatted as %)
            live_decl, proj["extrap_rate"], proj["same_as_last"],
            ("YES" if proj["extrap_rate"] is not None and r["threshold"] is not None
                  and proj["extrap_rate"] >= r["threshold"] else
             "NO" if proj["extrap_rate"] is not None and r["threshold"] is not None else ""),
            ("YES" if proj["same_as_last"] is not None and r["threshold"] is not None
                  and proj["same_as_last"] >= r["threshold"] else
             "NO" if proj["same_as_last"] is not None and r["threshold"] is not None else ""),
            edge,
        ]
        for i, v in enumerate(cells, 1):
            ws.cell(row=row, column=i, value=v)
        # Format numerics
        ws.cell(row=row, column=5).number_format = "$#,##0"
        ws.cell(row=row, column=6).number_format = "$#,##0"
        ws.cell(row=row, column=7).number_format = "0.000"
        ws.cell(row=row, column=8).number_format = "0.000"
        ws.cell(row=row, column=9).number_format = "0.0%"
        if isinstance(cells[10], (int, float)) and cells[10] is not None:
            ws.cell(row=row, column=11).number_format = "0"
        if isinstance(cells[11], (int, float)) and cells[11] is not None:
            ws.cell(row=row, column=12).number_format = "0"
        row += 1

    # Column widths
    widths = {1: 50, 2: 6, 3: 10, 4: 70, 5: 12, 6: 12, 7: 11, 8: 11, 9: 11,
              10: 14, 11: 18, 12: 18, 13: 14, 14: 14, 15: 60}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A6"

    # Conditional formatting on edge flag column (15) to highlight mismatches
    last_row = row - 1
    if last_row >= 6:
        rng = f"O6:O{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="notEqual", formula=['""'],
            fill=PatternFill("solid", fgColor="FFC7CE")))

    if not dry_run:
        save_with_retry(wb, XLSX)


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--watch", action="store_true")
    p.add_argument("--interval", type=int, default=60)
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    while True:
        try:
            rows = fetch_all_markets()
            live = get_live_data()
            write_markets_tab(rows, live, dry_run=args.dry_run)
            print(f"[{datetime.now():%H:%M:%S}] markets refreshed: {len(rows)} live contracts")
        except Exception as e:
            print(f"[{datetime.now():%H:%M:%S}] ERROR {e}", file=sys.stderr)
        if not args.watch:
            return
        time.sleep(args.interval)


if __name__ == "__main__":
    main()
