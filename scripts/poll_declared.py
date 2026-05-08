"""
Poll Democracy Club results API and write live DECLARED seats into the tracker.

Usage:
    python poll_declared.py            # one-shot update
    python poll_declared.py --watch    # poll every 5 minutes until stopped
    python poll_declared.py --watch --interval 120

Reads:  council_tracker_2026-05-08.xlsx (DECLARED block, cols 26-34, plus Status col F)
Writes: same file in place (close Excel before running)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import urllib.request
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

from xlsx_lock import workbook_lock, save_with_retry

ROOT = Path(__file__).parent
XLSX = ROOT / "council_tracker_2026-05-08.xlsx"

API = "https://candidates.democracyclub.org.uk/api/next/results/?election_date=2026-05-07&page_size=500"

# Party EC id -> our PARTIES index (Con, Lab, LD, Grn, Ref, SNP, PC, Ind, Oth)
PARTY_MAP = {
    "PP52": 0,                # Conservative and Unionist
    "PP53": 1,                # Labour
    "joint-party:53-119": 1,  # Labour & Co-operative
    "PP90": 2,                # Liberal Democrats
    "PP63": 3,                # Green Party (E&W)
    "PP1052": 3,              # Scottish Greens
    "PP7931": 4,              # Reform UK
    "PP102": 5,               # SNP
    "PP77": 6,                # Plaid Cymru
    "ynmp-party:2": 7,        # Independent
}

# Column layout: PRE H..P [8-16] | LAST Q..Y [17-25] | PREDICT Z..AH [26-34] | DECLARED AI..AQ [35-43]
DECL_START_COL = 35   # AI
N_PARTIES = 9
STATUS_COL = 6        # F
DECLARED_AT_COL = 7   # G
SEATS_UP_COL = 4      # D


def party_index(party_obj: dict) -> int:
    """Map an API party object -> PARTIES index. Unknown -> Oth (8)."""
    ec = party_obj.get("ec_id", "")
    if ec in PARTY_MAP:
        return PARTY_MAP[ec]
    name = (party_obj.get("name") or "").lower()
    # Heuristic fallbacks for joint/minor parties not in the explicit map
    if "labour" in name and "co-operative" in name:
        return 1
    if "conservative" in name:
        return 0
    if "labour" in name:
        return 1
    if "liberal democrat" in name:
        return 2
    if "green" in name:
        return 3
    if "reform" in name:
        return 4
    if "scottish national" in name or name == "snp":
        return 5
    if "plaid cymru" in name:
        return 6
    if "independent" in name:
        return 7
    return 8  # Oth


def slugify(name: str) -> str:
    """Workbook council name -> API-style slug for matching."""
    n = name
    n = re.sub(r"\s*\(Mayor\)\s*$", "", n)
    n = re.sub(r"\s*\(by-election\)\s*$", "", n)
    n = re.sub(r"\s*\(shadow\)\s*$", "", n)
    n = re.sub(r"\s*CC\s*$", "", n)
    n = n.replace("Hull (Kingston upon Hull)", "Kingston upon Hull")
    n = n.split(" — ")[0]  # for "Newport — Rogerstone North"
    n = n.replace("&", "and").replace("'", "").replace(".", "")
    n = re.sub(r"\s+", "-", n.strip()).lower()
    return n


# Slugs in the API that need translating to our slug form
SLUG_ALIASES = {
    "city-of-lincoln": "lincoln",
    "city-of-london": "city-of-london",  # not in our list, ignored
    "kingston-upon-hull-city": "kingston-upon-hull",
}


def fetch_all_results() -> list[dict]:
    out = []
    url = API
    while url:
        req = urllib.request.Request(url, headers={"User-Agent": "elections-tracker/1.0"})
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
        out.extend(data["results"])
        url = data.get("next")
        # urlopen can struggle with http: vs https inconsistencies in next-links
        if url and url.startswith("http://"):
            url = "https://" + url[len("http://"):]
    return out


def aggregate(results: list[dict]) -> dict[str, list[int]]:
    """council_slug -> [9 ints] elected counts."""
    out: dict[str, list[int]] = defaultdict(lambda: [0] * N_PARTIES)
    for r in results:
        bp = r["ballot"]["ballot_paper_id"]
        parts = bp.split(".")
        if parts[0] != "local":
            continue  # ignore parl/mayor/etc
        slug = parts[1]
        slug = SLUG_ALIASES.get(slug, slug)
        for cr in r["candidate_results"]:
            if cr.get("elected"):
                out[slug][party_index(cr["party"])] += 1
    return dict(out)


def update_workbook(declared: dict[str, list[int]], dry_run: bool = False) -> None:
    with workbook_lock():
        _update_workbook_inner(declared, dry_run)
    # Fetch BBC headline data (no workbook access; writes bbc_data.json)
    if not dry_run:
        try:
            from fetch_bbc import main as run_bbc_fetch
            run_bbc_fetch()
        except Exception as e:
            print(f"  BBC fetch failed: {e}", file=sys.stderr)
    # market_model takes its own lock — must run AFTER our lock is released
    if not dry_run:
        try:
            from market_model import run as run_market_model
            run_market_model()
        except Exception as e:
            print(f"  market_model failed: {e}", file=sys.stderr)
        # Then write Headline tab using Market model's calibrated totals
        try:
            with workbook_lock():
                wb2 = openpyxl.load_workbook(XLSX)
                update_headline_tab(wb2)
                save_with_retry(wb2, XLSX)
        except Exception as e:
            print(f"  headline tab failed: {e}", file=sys.stderr)


def _update_workbook_inner(declared: dict[str, list[int]], dry_run: bool) -> None:
    wb = openpyxl.load_workbook(XLSX)
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    changes = 0
    unmapped_in_workbook: list[str] = []

    # Build slug -> (sheet, row) index from the workbook
    slug_to_loc: dict[str, tuple[str, int]] = {}
    for sheet_name in ("England", "Scotland", "Wales"):
        ws = wb[sheet_name]
        for row in range(3, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            if not name or name == "TOTAL":
                continue
            slug_to_loc[slugify(name)] = (sheet_name, row)

    # Track which API slugs we couldn't place
    unmatched_api: list[str] = []
    for slug, counts in declared.items():
        if slug not in slug_to_loc:
            unmatched_api.append(slug)
            continue
        sheet_name, row = slug_to_loc[slug]
        ws = wb[sheet_name]
        # Read existing DECLARED to detect change
        existing = [ws.cell(row=row, column=DECL_START_COL + i).value or 0 for i in range(N_PARTIES)]
        if existing == counts:
            continue
        for i, v in enumerate(counts):
            ws.cell(row=row, column=DECL_START_COL + i, value=v)
        seats_up = ws.cell(row=row, column=SEATS_UP_COL).value or 0
        decl_sum = sum(counts)
        prev_status = ws.cell(row=row, column=STATUS_COL).value
        if decl_sum == 0:
            new_status = "Pending"
        elif decl_sum >= seats_up:
            new_status = "Complete"
        else:
            new_status = "Partial"
        # Don't downgrade a manually-set "Verified"
        if prev_status != "Verified":
            ws.cell(row=row, column=STATUS_COL, value=new_status)
        ws.cell(row=row, column=DECLARED_AT_COL, value=now)
        changes += 1

    council_name = lambda slug: slug_to_loc.get(slug, ("?", "?"))[0]
    print(f"[{now}] {len(declared)} councils with results | {changes} updated | "
          f"{len(unmatched_api)} api-slugs unmapped (sample: {unmatched_api[:5]})")

    update_projection_summary(wb)
    # Drop deprecated Per-council odds tab (replaced by Market model tab,
    # which is written by market_model.run() called from update_workbook).
    if "Per-council odds" in wb.sheetnames:
        del wb["Per-council odds"]
    if not dry_run:
        save_with_retry(wb, XLSX)


PARTIES = ["Con", "Lab", "LD", "Grn", "Ref", "SNP", "PC", "Ind", "Oth"]
PRE_C, LAST_C, PREDICT_C, DECL_C = 8, 17, 26, 35  # column-block starts


def update_projection_summary(wb) -> None:
    """Compute per-party live projections in Python, write static values to Summary tab.

    Three projection methods, all = (declared real) + (model for remaining):
      Same-as-last : remaining = last_party * remaining_seats / sum(last)
                      -> assumes remaining councils repeat 2022 pattern (naive)
      Swing-adjusted: remaining = max(0, last_share + national_swing) renormalised * remaining_seats
                      -> applies observed live swing; far more realistic mid-count
      PREDICT-based : remaining = predict_party * remaining_seats / sum(predict)
                      -> falls back to swing-adjusted when PREDICT block is empty for a council
    """
    swing = compute_national_swing(wb)
    pc_path = Path(__file__).parent / "pollcheck.json"
    pollcheck = json.loads(pc_path.read_text(encoding="utf-8")) if pc_path.exists() else {}

    decl = [0] * 9
    proj_last = [0.0] * 9
    proj_swing = [0.0] * 9
    proj_pred = [0.0] * 9
    proj_pc = [0.0] * 9   # PollCheck-hybrid: declared (if Complete) else PollCheck central

    for sheet_name in ("England", "Wales"):
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if not name or name == "TOTAL":
                continue
            seats_up = ws.cell(row=r, column=SEATS_UP_COL).value or 0
            if not seats_up:
                continue
            status = ws.cell(row=r, column=STATUS_COL).value
            last_arr = [ws.cell(row=r, column=LAST_C + i).value or 0 for i in range(9)]
            pred_arr = [ws.cell(row=r, column=PREDICT_C + i).value or 0 for i in range(9)]
            decl_arr = [ws.cell(row=r, column=DECL_C + i).value or 0 for i in range(9)]
            d_sum = sum(decl_arr)
            remaining = max(seats_up - d_sum, 0)
            l_sum = sum(last_arr)
            p_sum = sum(pred_arr)

            # Swing-adjusted shares for THIS council
            if l_sum:
                shifted = [max(0.0, last_arr[i] / l_sum + swing[i]) for i in range(9)]
                ssum = sum(shifted) or 1.0
                swing_share = [s / ssum for s in shifted]
            else:
                swing_share = [0.0] * 9

            # PollCheck-hybrid contribution
            pc = pollcheck.get(name)
            if status in ("Complete", "Verified") and d_sum > 0:
                pc_contrib = decl_arr[:]
            elif pc and pc.get("projection"):
                pc_contrib = [0] * 9
                pc_party_idx = {p: i for i, p in enumerate(PARTIES)}
                for row_proj in pc["projection"]:
                    idx = pc_party_idx.get(row_proj["party"], 8)
                    pc_contrib[idx] += row_proj["central"]
                pc_total = sum(pc_contrib)
                if pc_total > 0 and pc_total != seats_up:
                    pc_contrib = [v * seats_up / pc_total for v in pc_contrib]
            else:
                pc_contrib = [decl_arr[i] + swing_share[i] * remaining for i in range(9)]

            for i in range(9):
                decl[i] += decl_arr[i]
                proj_last[i]  += decl_arr[i] + (last_arr[i] * remaining / l_sum if l_sum else 0)
                proj_swing[i] += decl_arr[i] + swing_share[i] * remaining
                proj_pc[i]    += pc_contrib[i]
                if p_sum:
                    proj_pred[i] += decl_arr[i] + pred_arr[i] * remaining / p_sum
                else:
                    proj_pred[i] += decl_arr[i] + swing_share[i] * remaining

    decl_total = sum(decl)
    last_total = sum(proj_last)
    swing_total = sum(proj_swing)
    pred_total = sum(proj_pred)
    pc_total = sum(proj_pc)

    # Locate projection block on Summary tab
    sm = wb["Summary"]
    proj_row = None
    for r in range(1, sm.max_row + 1):
        v = sm.cell(row=r, column=1).value
        if v and "Live projection" in str(v):
            proj_row = r
            break
    if not proj_row:
        return

    # Rewrite headers (5 projections side by side)
    new_headers = ["Party", "Declared", "% of declared",
                   "Same-as-last total", "% of total",
                   "Swing-adjusted total", "% of total",
                   "PollCheck-hybrid total", "% of total",
                   "PREDICT-based total", "% of total"]
    from openpyxl.styles import Font, PatternFill
    for i, h in enumerate(new_headers, 1):
        c = sm.cell(row=proj_row + 1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    for i, party in enumerate(PARTIES):
        r = proj_row + 2 + i
        sm.cell(row=r, column=2, value=decl[i])
        sm.cell(row=r, column=3, value=(decl[i] / decl_total) if decl_total else 0)
        sm.cell(row=r, column=3).number_format = "0.0%"
        sm.cell(row=r, column=4, value=round(proj_last[i]))
        sm.cell(row=r, column=5, value=(proj_last[i] / last_total) if last_total else 0)
        sm.cell(row=r, column=5).number_format = "0.0%"
        sm.cell(row=r, column=6, value=round(proj_swing[i]))
        sm.cell(row=r, column=7, value=(proj_swing[i] / swing_total) if swing_total else 0)
        sm.cell(row=r, column=7).number_format = "0.0%"
        sm.cell(row=r, column=8, value=round(proj_pc[i]))
        sm.cell(row=r, column=9, value=(proj_pc[i] / pc_total) if pc_total else 0)
        sm.cell(row=r, column=9).number_format = "0.0%"
        sm.cell(row=r, column=10, value=round(proj_pred[i]))
        sm.cell(row=r, column=11, value=(proj_pred[i] / pred_total) if pred_total else 0)
        sm.cell(row=r, column=11).number_format = "0.0%"

    ptot = proj_row + 2 + len(PARTIES)
    sm.cell(row=ptot, column=1, value="TOTAL").font = Font(bold=True)
    sm.cell(row=ptot, column=2, value=decl_total).font = Font(bold=True)
    sm.cell(row=ptot, column=4, value=round(last_total)).font = Font(bold=True)
    sm.cell(row=ptot, column=6, value=round(swing_total)).font = Font(bold=True)
    sm.cell(row=ptot, column=8, value=round(pc_total)).font = Font(bold=True)
    sm.cell(row=ptot, column=10, value=round(pred_total)).font = Font(bold=True)


def update_headline_tab(wb) -> None:
    """Single one-page headline view: declared seats, predicted seats, conclusions.
       Always positioned as the first tab in the workbook."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    # Pull declared totals from BBC (faster source) if available, else from DC region tabs
    decl = [0] * 9
    seats_up_total = 0
    councils_total = 0
    councils_complete = 0
    bbc_path = Path(__file__).parent / "bbc_data.json"
    bbc_data = json.loads(bbc_path.read_text(encoding="utf-8")) if bbc_path.exists() else None

    for sheet_name in ("England", "Wales"):
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if not name or name == "TOTAL":
                continue
            seats_up = ws.cell(row=r, column=SEATS_UP_COL).value or 0
            if not seats_up:
                continue
            councils_total += 1
            seats_up_total += seats_up
            if ws.cell(row=r, column=STATUS_COL).value in ("Complete", "Verified"):
                councils_complete += 1

    if bbc_data:
        seats_map = bbc_data.get("seats_by_party", {})
        for i, p in enumerate(PARTIES):
            decl[i] = seats_map.get(p, 0) or 0
        bbc_councils_declared = bbc_data.get("councils_declared") or 0
        bbc_councils_total = bbc_data.get("councils_total") or councils_total
        # Use BBC's councils declared as the headline (more current)
        councils_complete = bbc_councils_declared
        # Note: BBC's councils_total (136) excludes 2 shadow Surrey councils + Newport ward;
        # we can keep our 144 as authoritative for "councils tracked", but display BBC's count
    else:
        # Fallback to DC API data already in region tabs
        for sheet_name in ("England", "Wales"):
            ws = wb[sheet_name]
            for r in range(3, ws.max_row + 1):
                name = ws.cell(row=r, column=1).value
                if not name or name == "TOTAL":
                    continue
                seats_up = ws.cell(row=r, column=SEATS_UP_COL).value or 0
                if not seats_up:
                    continue
                for i in range(9):
                    decl[i] += ws.cell(row=r, column=DECL_C + i).value or 0

    # Pull calibrated predicted totals from Market model tab's aggregate row.
    # Falls back to per-council PollCheck-hybrid sum if Market model not yet written.
    proj = [0.0] * 9
    if "Market model" in wb.sheetnames:
        mm = wb["Market model"]
        for r in range(6, mm.max_row + 1):
            cell = mm.cell(row=r, column=1).value
            if cell and "Aggregate expected seats" in str(cell):
                # Aggregate row spans cols 21..29 (after BBC cols added)
                for i in range(9):
                    v = mm.cell(row=r, column=21 + i).value
                    proj[i] = float(v) if isinstance(v, (int, float)) else 0.0
                break
    if sum(proj) < 100:  # market model not present or empty - fall back to declared as crude proxy
        proj = [float(v) for v in decl]

    decl_total = sum(decl) or 1
    proj_total = sum(proj) or 1

    # Build the tab
    sheet_name = "Headline"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, 0)  # first tab

    BLUE = PatternFill("solid", fgColor="1F4E78")
    LIGHT_BLUE = PatternFill("solid", fgColor="D9E1F2")
    GREY = PatternFill("solid", fgColor="F2F2F2")
    WHITE_BOLD = Font(bold=True, color="FFFFFF", size=12)
    THIN = Side(border_style="thin", color="999999")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    # Title row
    ws["A1"] = "UK Council Elections 2026 — Live Headline"
    ws["A1"].font = Font(bold=True, size=18)
    ws.merge_cells("A1:E1")

    src = "BBC News" if bbc_data else "Democracy Club API"
    ws["A2"] = (f"Updated {datetime.now():%H:%M:%S}  |  Declared data: {src}  |  "
                f"{councils_complete} of {councils_total} councils declared "
                f"({100*councils_complete/councils_total:.0f}%)  |  "
                f"{int(decl_total):,} of {seats_up_total:,} seats counted "
                f"({100*decl_total/seats_up_total:.1f}%)")
    ws["A2"].font = Font(italic=True, color="666666", size=11)
    ws.merge_cells("A2:E2")

    # Headers row 4
    headers = ["Party", "Declared seats", "% of declared", "Predicted final", "% of final"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.font = WHITE_BOLD
        c.fill = BLUE
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    # Sort parties: Reform first if leading, otherwise by predicted seats descending
    party_order = sorted(range(9), key=lambda i: -proj[i])
    party_order = [i for i in party_order if proj[i] > 0 or decl[i] > 0]

    PARTY_FULL = {"Con": "Conservative", "Lab": "Labour", "LD": "Liberal Democrats",
                  "Grn": "Green", "Ref": "Reform UK", "SNP": "SNP", "PC": "Plaid Cymru",
                  "Ind": "Independent", "Oth": "Other"}

    PARTY_FILL = {
        "Con": "0087DC",  # Tory blue
        "Lab": "E4003B",  # Labour red
        "LD": "FAA61A",   # LD orange
        "Grn": "6AB023",  # Green
        "Ref": "12B6CF",  # Reform teal
        "SNP": "FDF38E",  # SNP yellow
        "PC": "008142",   # Plaid green
        "Ind": "BFBFBF",  # grey
        "Oth": "D9D9D9",  # light grey
    }

    row = 5
    for i in party_order:
        party = PARTIES[i]
        ws.cell(row=row, column=1, value=PARTY_FULL[party]).font = Font(bold=True, size=11)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=PARTY_FILL[party])
        ws.cell(row=row, column=2, value=int(decl[i]))
        ws.cell(row=row, column=3, value=decl[i] / decl_total).number_format = "0.0%"
        ws.cell(row=row, column=4, value=int(round(proj[i])))
        ws.cell(row=row, column=5, value=proj[i] / proj_total).number_format = "0.0%"
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = BORDER
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).fill = GREY
    ws.cell(row=row, column=2, value=int(decl_total)).font = Font(bold=True)
    ws.cell(row=row, column=4, value=int(round(proj_total))).font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=row, column=col).border = BORDER
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
    row += 2

    # "What this means" section
    winner_idx = party_order[0]
    runner_idx = party_order[1] if len(party_order) > 1 else winner_idx
    third_idx = party_order[2] if len(party_order) > 2 else winner_idx

    winner_name = PARTY_FULL[PARTIES[winner_idx]]
    runner_name = PARTY_FULL[PARTIES[runner_idx]]
    third_name = PARTY_FULL[PARTIES[third_idx]]
    winner_seats = int(round(proj[winner_idx]))
    runner_seats = int(round(proj[runner_idx]))
    third_seats = int(round(proj[third_idx]))
    margin = winner_seats - runner_seats

    ws.cell(row=row, column=1, value="What this means").font = WHITE_BOLD
    ws.cell(row=row, column=1).fill = BLUE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1

    bullets = [
        f"{winner_name} on track to win the most seats — projected {winner_seats:,} of {int(round(proj_total)):,} ({100*proj[winner_idx]/proj_total:.1f}%).",
        f"{runner_name} projected second with {runner_seats:,} seats ({100*proj[runner_idx]/proj_total:.1f}%) — {margin:,}-seat margin behind {winner_name}.",
        f"{third_name} third on {third_seats:,} ({100*proj[third_idx]/proj_total:.1f}%).",
        f"Counting is {100*decl_total/seats_up_total:.1f}% complete; remaining {seats_up_total - int(decl_total):,} seats may shift these numbers.",
        "Predicted column = Market model: PollCheck per-council projections (FPTP-aware) calibrated to Polymarket aggregate, plus actual declared seats.",
        "Live betting market (Polymarket): " + _market_summary(wb),
    ]
    for b in bullets:
        ws.cell(row=row, column=1, value="•  " + b).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.row_dimensions[row].height = 22
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14


def _market_summary(wb) -> str:
    """Pull headline market prices from the Markets tab if present."""
    if "Markets" not in wb.sheetnames:
        return "(Markets tab not yet populated)"
    ws = wb["Markets"]
    bits = []
    for r in range(6, ws.max_row + 1):
        q = ws.cell(row=r, column=4).value or ""
        yes = ws.cell(row=r, column=7).value
        if not isinstance(yes, (int, float)):
            continue
        if "win the most" in q and "Reform UK" in q:
            bits.append(f"Reform 1st {yes*100:.0f}%")
        elif "second-most" in q and "Labour" in q:
            bits.append(f"Lab 2nd {yes*100:.0f}%")
        elif "Reform UK win at least 1600" in q:
            bits.append(f"Reform 1600+ {yes*100:.0f}%")
        elif "Reform UK win at least 1800" in q:
            bits.append(f"Reform 1800+ {yes*100:.0f}%")
    return "  ·  ".join(bits) if bits else "(no market prices loaded)"


def compute_national_swing(wb) -> list[float]:
    """Size-weighted swing in seat-share terms across Complete councils.
       swing[i] = (sum_decl[i] / total_decl) - (sum_last[i] / total_last)
       This weights big councils (Birmingham 101) appropriately vs tiny districts."""
    tot_decl = [0] * 9
    tot_last = [0] * 9
    for sheet_name in ("England", "Wales"):
        ws = wb[sheet_name]
        for r in range(3, ws.max_row + 1):
            name = ws.cell(row=r, column=1).value
            if not name or name == "TOTAL":
                continue
            status = ws.cell(row=r, column=STATUS_COL).value
            if status not in ("Complete", "Verified"):
                continue
            seats_up = ws.cell(row=r, column=SEATS_UP_COL).value or 0
            if not seats_up:
                continue
            last_arr = [ws.cell(row=r, column=LAST_C + i).value or 0 for i in range(9)]
            decl_arr = [ws.cell(row=r, column=DECL_C + i).value or 0 for i in range(9)]
            l_sum = sum(last_arr); d_sum = sum(decl_arr)
            if not l_sum or not d_sum:
                continue
            for i in range(9):
                tot_decl[i] += decl_arr[i]
                tot_last[i] += last_arr[i]
    sd = sum(tot_decl); sl = sum(tot_last)
    if not sd or not sl:
        return [0.0] * 9
    return [tot_decl[i] / sd - tot_last[i] / sl for i in range(9)]


def update_per_council_odds(wb) -> None:
    """Write a 'Per-council odds' tab.

    For each council:
      - Complete -> winner = leading party in DECLARED, prob = 1.0
      - Pending/Partial -> apply national swing (computed from Complete councils)
        to that council's last-election shares; pick argmax as expected winner;
        derive a probability from margin-to-2nd via logistic.
    """
    swing = compute_national_swing(wb)

    # Load PollCheck cache if present
    pc_path = Path(__file__).parent / "pollcheck.json"
    pollcheck = json.loads(pc_path.read_text(encoding="utf-8")) if pc_path.exists() else {}

    sheet_name = "Per-council odds"
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ws["A1"] = "Per-council live-swing model — winner probabilities"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"National swing observed (Complete councils): "
                + " ".join(f"{p}{'+' if s>=0 else ''}{s*100:.1f}pp" for p, s in zip(PARTIES, swing) if abs(s) > 0.005))
    ws["A2"].font = Font(italic=True, color="666666")
    ws["A3"] = ("Pending/Partial councils: party share = max(0, last_share + swing); "
                "winner = argmax of projected seats; prob = logistic(margin/seats_up).")
    ws["A3"].font = Font(italic=True, color="666666")

    headers = ["Region", "Council", "Type", "Sub-region", "Status", "Seats up",
               "Modelled winner", "Win prob", "Runner-up", "Margin",
               "PollCheck winner", "PollCheck prob"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.alignment = Alignment(horizontal="center")

    import math

    def logistic(x: float, k: float = 8.0) -> float:
        # x is margin/seats_up in [0,1]. k tuned so a 25% margin -> ~88% prob.
        return 1 / (1 + math.exp(-k * x))

    row = 6
    for region in ("England", "Wales", "Scotland"):
        rs = wb[region]
        for r in range(3, rs.max_row + 1):
            name = rs.cell(row=r, column=1).value
            if not name or name == "TOTAL":
                continue
            ctype = rs.cell(row=r, column=2).value
            subr = rs.cell(row=r, column=3).value
            seats_up = rs.cell(row=r, column=SEATS_UP_COL).value or 0
            status = rs.cell(row=r, column=STATUS_COL).value
            last_arr = [rs.cell(row=r, column=LAST_C + i).value or 0 for i in range(9)]
            decl_arr = [rs.cell(row=r, column=DECL_C + i).value or 0 for i in range(9)]

            if status in ("Complete", "Verified") and sum(decl_arr) > 0:
                # Use actual declared
                seats = decl_arr[:]
                d_sum = sum(seats)
                shares = [s / d_sum for s in seats] if d_sum else [0]*9
                winner_i = max(range(9), key=lambda i: seats[i])
                second_i = sorted(range(9), key=lambda i: seats[i])[-2]
                prob = 1.0
                margin = seats[winner_i] - seats[second_i]
            else:
                # Project: last-share + swing, clipped >=0
                l_sum = sum(last_arr)
                if not l_sum or not seats_up:
                    continue
                proj_share = [max(0.0, last_arr[i] / l_sum + swing[i]) for i in range(9)]
                ssum = sum(proj_share) or 1.0
                proj_share = [s / ssum for s in proj_share]  # renormalise
                proj_seats = [round(s * seats_up) for s in proj_share]
                # If we have partial declared, blend: declared + projected for remaining
                d_sum = sum(decl_arr)
                if status == "Partial" and d_sum > 0:
                    remaining = max(seats_up - d_sum, 0)
                    proj_seats = [decl_arr[i] + round(proj_share[i] * remaining) for i in range(9)]
                seats = proj_seats
                ranking = sorted(range(9), key=lambda i: seats[i], reverse=True)
                winner_i = ranking[0]
                second_i = ranking[1] if len(ranking) > 1 else winner_i
                margin = seats[winner_i] - seats[second_i]
                prob = logistic(margin / seats_up) if seats_up else 0.5

            ws.cell(row=row, column=1, value=region)
            ws.cell(row=row, column=2, value=name)
            ws.cell(row=row, column=3, value=ctype)
            ws.cell(row=row, column=4, value=subr)
            ws.cell(row=row, column=5, value=status or "Pending")
            ws.cell(row=row, column=6, value=seats_up)
            ws.cell(row=row, column=7, value=PARTIES[winner_i])
            ws.cell(row=row, column=8, value=prob).number_format = "0.0%"
            ws.cell(row=row, column=9, value=PARTIES[second_i])
            ws.cell(row=row, column=10, value=margin)

            # PollCheck columns
            pc = pollcheck.get(name)
            if pc and pc.get("projection"):
                # Monte Carlo on uniform [low, high] per party -> P(party has most seats)
                import random
                rows_pc = pc["projection"]
                trials = 4000
                wins = {r["party"]: 0 for r in rows_pc}
                for _ in range(trials):
                    samples = [(r["party"], random.uniform(r["low"], r["high"])) for r in rows_pc]
                    winner = max(samples, key=lambda x: x[1])[0]
                    wins[winner] = wins.get(winner, 0) + 1
                top_party, top_count = max(wins.items(), key=lambda kv: kv[1])
                ws.cell(row=row, column=11, value=top_party)
                ws.cell(row=row, column=12, value=top_count / trials).number_format = "0.0%"

            row += 1

    widths = {1: 11, 2: 36, 3: 14, 4: 18, 5: 12, 6: 9, 7: 14, 8: 10,
              9: 12, 10: 9, 11: 16, 12: 12}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A6"


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--watch", action="store_true", help="loop forever")
    p.add_argument("--interval", type=int, default=300, help="seconds between polls (default 300)")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    while True:
        try:
            results = fetch_all_results()
            agg = aggregate(results)
            update_workbook(agg, dry_run=args.dry_run)
        except Exception as e:
            print(f"[{datetime.now():%H:%M:%S}] ERROR {e}", file=sys.stderr)
        if not args.watch:
            return
        time.sleep(args.interval)


if __name__ == "__main__":
    main()
